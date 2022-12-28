#!/usr/bin/env python3
#
# Copyright (c) 2021, Pelion Limited and affiliates.
# Copyright (c) 2022 Izuma Networks
#
# SPDX-License-Identifier: Apache-2.0
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.


"""
A tool for dealing with Yocto license manifest files.

This tool can converting them to CSV/Excel-format,
optionally with change information.
"""


import sys
import os
import logging
import argparse
import re
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# Lots of literals
_CSV = ".csv"
_XLS = ".xlsx"
_EXISTS = "'.csv or .xlsx already exists."
_NOT_EXIST = "' does not exist."
_OVERWRITE = " Will overwrite."

_PREV_LIC = "Previous license"
_PREV_VER = "Previous version"
_PREV_REC = "Previous recipe"
_CURR_LIC = "Current license"
_CURR_VER = "Current version"
_CURR_REC = "Current recipe"

_MARK_CHG = "y"  # What we use to mark changes
_PKG = "Package"
_CHG = "Change"
_PKG_ADD = "Package added"
_PKG_REM = "Package removed"
_LIC_CHG = "License change"
_VER_CHG = "Version change"

_DATA_SHEET_NAME = "sheet1"


def _print_help():

    print("Yocto license manifest tool")
    print("")
    print("license-tool.py list <input license manifest file> <output file>")
    print("   Generate a CSV-formatted version of the Yocto manifest file")
    print("   For example license-tool.py list license.manifest licenses ")
    print("   Will generate licenses.csv and licenses.xlsx files.")
    print("")
    print(
        "license-tool.py changes <previous manifest file>"
        " <current manifest file> <output file>"
    )
    print(
        "   Generate a CSV and Excel-formatted version based on two Yocto"
        " manifest files that highlights the changes"
    )
    print(
        "   For example license-tool.py changes license.manifest.v82"
        " license.manifest.v83 license-changes-v82-v83"
    )
    print(
        "   Will generate license-changes-v82-v83.csv and "
        " license-changes-v82-v83.xlsx files."
    )
    print("")
    print("Optional:")
    print(" --verbose   verbose output")
    print(" --debug     debug level output")
    print(" --force     enable overwriting of existing output files")


# read_manifest_file - read Yocto license manifest and turn into a Panda's
#                      dataframe and a status dictionary.
#
def read_manifest_file(input_file):
    """Read manifest file and return a Panda's dataframe ."""
    pattern = re.compile(
        "PACKAGE NAME: (.*)\nPACKAGE VERSION: (.*)\nRECIPE NAME: (.*)\n"
        "LICENSE: (.*)\n\n"
    )
    with open(input_file, "r", encoding="utf8") as f_h:
        data = f_h.read()
        # Create empty dataframe
        column_names = [_PKG, "version", "recipe", "license"]
        d_f = pd.DataFrame(columns=column_names)

        package_count = 0
        errors = False
        prew = 0
        for info_field in re.finditer(pattern, data):
            package_count += 1
            if info_field.span()[0] != prew:
                print(
                    "ERROR - Invalid content in the file, got "
                    + str(package_count)
                    + " packages."
                )
                # There is some content not matching the pattern in file.
                errors = True
            prew = info_field.span()[1]

            new_row = {
                column_names[0]: info_field.group(1),
                column_names[1]: info_field.group(2),
                column_names[2]: info_field.group(3),
                column_names[3]: info_field.group(4),
            }
            d_f = d_f.append(new_row, ignore_index=True)

        if (
            package_count == 0
        ):  # needs to have at least one package or it is an error
            print("Package count is zero")
            errors = True

        data_len = len(data)
        if data_len != prew:
            # if not all data was matched it is an error
            print(
                "ERROR - Invalid content at end of file, got "
                + str(package_count)
                + " packages."
            )
            errors = True

        num_lines = sum(1 for line in data.split("\n"))
        status = {
            "lines": num_lines,
            "packages": package_count,
            "errors": errors,
        }
    return d_f, status


# gen_list - generate list-formatted files from a license manifest file
#            filenames of input file and output filename base needed as
#            parameters. Two output files are created, file.csv and .xlsx.
#
def gen_list(inputfile, outputfile):
    """
    gen_list - generate list.

    generate list formatted output (.csv and .xlsx) from
    manifest file.
    """
    logging.debug("gen_list: '%s','%s'", inputfile, outputfile)

    # Covert the manifest to a Pandas dataframe
    d_f, status = read_manifest_file(inputfile)

    # Export CSV, if no errors noticed
    if status["errors"] is False:
        d_f.to_csv(outputfile + _CSV, index=False)
        generate_excel(
            outputfile + _XLS,
            d_f.style,
            template_file="excel-template-list.xlsx",
        )
    else:
        print("ERROR - could not process license manifest file " + inputfile)
        sys.exit(71)  # EPROTO
    print(inputfile + " " + str(status))


# init_change_summary  - initialize change summary dict
#
def init_change_summary():
    """Init and return a change_summary dict."""
    logging.info("Finding changes...")
    change_summary = {
        _VER_CHG: 0,
        _LIC_CHG: 0,
        _PKG_ADD: 0,
        _PKG_REM: 0,
    }
    return change_summary


#
# print_change_summary - print out a summary of the changes
#                        from the dict passed as argumntn (chg_sum)
#
def print_change_summary(chg_sum):
    """Print change summary from given dict."""
    print("")
    print(
        "Total changes  : ",
        (
            chg_sum[_PKG_ADD]
            + chg_sum[_PKG_REM]
            + chg_sum[_LIC_CHG]
            + chg_sum[_VER_CHG]
        ),
    )
    print("Package changes: ", (chg_sum[_PKG_ADD] + chg_sum[_PKG_REM]))
    print("- Added        : ", chg_sum[_PKG_ADD])
    print("- Removed      : ", chg_sum[_PKG_REM])
    print("License changes: ", chg_sum[_LIC_CHG])
    print("Version changes: ", chg_sum[_VER_CHG])


# read_and_merge_manifests  - read previous and current manifest files
#                             and return a merged dataframe with their content.


def read_and_merge_manifests(previous, current):
    """Read previous and current manifest, return as two dataframes."""
    d_f_prev, status_prev = read_manifest_file(previous)
    d_f_prev.rename(
        columns={
            "version": _PREV_VER,
            "license": _PREV_LIC,
            "recipe": _PREV_REC,
        },
        inplace=True,
    )
    if status_prev["errors"] is True:
        print("ERROR - handling of '" + previous + "' failed.")
        sys.exit(71)  # EPROTO
    print(previous + " " + str(status_prev))

    d_f_curr, status_curr = read_manifest_file(current)
    d_f_curr.rename(
        columns={
            "version": _CURR_VER,
            "license": _CURR_LIC,
            "recipe": _CURR_REC,
        },
        inplace=True,
    )
    if status_curr["errors"] is True:
        print("ERROR - handling of '" + current + "' failed.")
        sys.exit(71)  # EPROTO
    print(current + " " + str(status_curr))

    # 1st create a merged table that has both previous and current information
    logging.info("Merging dataframes...")
    d_f_combo = pd.merge(d_f_prev, d_f_curr, on=_PKG, how="outer")
    logging.debug(d_f_combo)
    return d_f_combo


# gen_changes - generate change information based on two Yocto
#               license manifest files
#
def gen_changes(previous, current, output):
    """
    gen_list - generate list.

    Generate list formatted change output (.csv and .xlsx)
    from 2 manifest files.
    """
    logging.debug("gen_changes: '%s', '%s', '%s'", previous, current, output)
    # Read the manifests, get them merged into one combined dataframe
    d_f_combo = read_and_merge_manifests(previous, current)

    # With that, we can now start going it through and add the "changes"
    # columns to highlight what has changed.
    # Not going to consider recipe change a change worth high-lighting,
    # that is not relevant from Third Party IP point of view.

    d_f_combo[[_CHG, _VER_CHG, _LIC_CHG, _PKG_ADD, _PKG_REM]] = ""
    i = 0
    rows = d_f_combo.shape[0]
    styled = d_f_combo.style
    change_summary = init_change_summary()  # Get dict for collecting changes
    while i < rows:
        # Check package appearing, start by setting change is "n"
        package_change = False
        if pd.isna(d_f_combo.at[i, _PREV_REC]):  # NaN
            d_f_combo.at[i, _CHG] = _MARK_CHG
            d_f_combo.at[i, _PKG_ADD] = _MARK_CHG
            styled = styled.apply(
                style_single_cell,
                row=i,
                column=d_f_combo.columns.get_loc(_CURR_REC),
                color="yellow",
                axis=None,
            )
            styled = styled.apply(
                style_single_cell,
                row=i,
                column=d_f_combo.columns.get_loc(_PKG),
                color="yellow",
                axis=None,
            )
            styled = styled.apply(
                style_single_cell,
                row=i,
                column=d_f_combo.columns.get_loc(_CURR_VER),
                color="yellow",
                axis=None,
            )
            styled = styled.apply(
                style_single_cell,
                row=i,
                column=d_f_combo.columns.get_loc(_CURR_LIC),
                color="yellow",
                axis=None,
            )
            package_change = True
            change_summary[_PKG_ADD] += 1
        # Package removed
        if package_change is False and pd.isna(
            d_f_combo.at[i, _CURR_REC]
        ):  # NaN
            d_f_combo.at[i, _CHG] = _MARK_CHG
            d_f_combo.at[i, _PKG_REM] = _MARK_CHG
            styled = styled.apply(
                style_single_cell,
                row=i,
                column=d_f_combo.columns.get_loc(_PREV_REC),
                color="yellow",
                axis=None,
            )
            styled = styled.apply(
                style_single_cell,
                row=i,
                column=d_f_combo.columns.get_loc(_PREV_VER),
                color="yellow",
                axis=None,
            )
            styled = styled.apply(
                style_single_cell,
                row=i,
                column=d_f_combo.columns.get_loc(_PKG),
                color="yellow",
                axis=None,
            )
            styled = styled.apply(
                style_single_cell,
                row=i,
                column=d_f_combo.columns.get_loc(_PREV_LIC),
                color="yellow",
                axis=None,
            )
            package_change = True
            change_summary[_PKG_REM] += 1
        # Version change
        if (
            package_change is False
            and d_f_combo.at[i, _PREV_VER] != d_f_combo.at[i, _CURR_VER]
        ):
            d_f_combo.at[i, _CHG] = _MARK_CHG
            d_f_combo.at[i, _VER_CHG] = _MARK_CHG
            styled = styled.apply(
                style_single_cell,
                row=i,
                column=d_f_combo.columns.get_loc(_PREV_VER),
                color="green",
                axis=None,
            )
            styled = styled.apply(
                style_single_cell,
                row=i,
                column=d_f_combo.columns.get_loc(_CURR_VER),
                color="green",
                axis=None,
            )
            change_summary[_VER_CHG] += 1
        # License change
        if (
            package_change is False
            and d_f_combo.at[i, _PREV_LIC] != d_f_combo.at[i, _CURR_LIC]
        ):
            d_f_combo.at[i, _CHG] = _MARK_CHG
            d_f_combo.at[i, _LIC_CHG] = _MARK_CHG
            styled = styled.apply(
                style_single_cell,
                row=i,
                column=d_f_combo.columns.get_loc(_PREV_LIC),
                color="red",
                axis=None,
            )
            styled = styled.apply(
                style_single_cell,
                row=i,
                column=d_f_combo.columns.get_loc(_CURR_LIC),
                color="red",
                axis=None,
            )
            change_summary[_LIC_CHG] += 1
        # No changes cases is the default, as we set all
        # change columns to n at start
        i = i + 1
    # Export result out
    logging.info("Export CSV: %s ", output + _CSV)
    d_f_combo.to_csv(path_or_buf=output + _CSV, index=False)
    logging.info("Export Excel: %s", output + _XLS)
    generate_excel(
        output=output + _XLS,
        styled=styled,
        template_file="excel-template-changes.xlsx",
    )
    print_change_summary(change_summary)


#
# generate_excel   output = output filename,
#                  styled = styled Pandas dataframe
#
def generate_excel(output, styled, template_file=None):
    """Generate Excel-file (output) from styled Panda's dataframe."""
    # Add autofilters to Excel sheet
    # pylint: disable=abstract-class-instantiated
    writer = pd.ExcelWriter(output, engine="openpyxl")

    if template_file:
        template_book = load_workbook(template_file)
        writer.book = template_book

    styled.to_excel(writer, sheet_name=_DATA_SHEET_NAME, index=False)

    # Get the xlsxwriter workbook and worksheet objects.
    # pylint: disable=E1101
    workbook = writer.book
    worksheet = workbook[_DATA_SHEET_NAME]

    # put the datasheet first, _sheets is protected
    # pylint: disable=W0212
    oldindex = workbook._sheets.index(worksheet)
    # pylint: disable=W0212
    workbook._sheets.pop(oldindex)
    # pylint: disable=W0212
    workbook._sheets.insert(0, worksheet)

    worksheet.auto_filter.ref = worksheet.dimensions

    colum_names = []
    for cell in worksheet[1]:
        colum_names.append(str(cell.value))

    # set default width of colums to match the title
    for col in range(worksheet.min_column, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(col)].width = (
            len(colum_names[col - 1]) + 5
        )

    writer.save()


#
# style_single_cell - set background color on a cell
#
def style_single_cell(work_sheet, row, column, color):
    """Set a single cell to a specific color."""
    new_color = f"background-color:  {color}"
    df1 = pd.DataFrame("", index=work_sheet.index, columns=work_sheet.columns)
    df1.iloc[row, column] = new_color
    return df1


def _parse_args():

    parser = argparse.ArgumentParser()

    # Two command modes, list and change

    subparsers = parser.add_subparsers(dest="command")
    parser_list = subparsers.add_parser(
        "list",
        help="Create CSV and Excel-file of Yocto License manifest file.",
    )
    parser_list.add_argument(
        "inputfile",
        help="Yocto license manifest file name (used as input)",
    )
    parser_list.add_argument(
        "listfile",
        help="Name base for CSV/Excel-formatted output list file name, i.e."
        " <listfile>.csv/.xlsx",
    )
    parser_changes = subparsers.add_parser(
        "changes",
        help="Create changes highlighting list from two Yocto license "
        "manifest files.",
    )
    parser_changes.add_argument(
        "previous",
        help="Previous Yocto license manifest file name",
    )
    parser_changes.add_argument(
        "current",
        help="Current Yocto license manifest file name",
    )
    parser_changes.add_argument(
        "changefile",
        help="Name base for CSV/Excel-formatted changes file name, i.e."
        " <changefile>.csv/.xlsx",
    )
    parser.add_argument(
        "--force",
        help="Force overwrite of existing output file",
        action="store_true",
    )
    parser.add_argument(
        "--verbose",
        help="Verbose diagnostics",
        action="store_const",
        dest="loglevel",
        const=logging.INFO,
    )
    parser.add_argument(
        "--debug",
        help="Extra diagnostic output",
        action="store_const",
        dest="loglevel",
        const=logging.DEBUG,
        default=logging.WARNING,
    )
    args, unknown = parser.parse_known_args()
    logging.basicConfig(level=args.loglevel, format="")
    if not args.command:
        _print_help()
        sys.exit(0)

    if len(unknown) > 0:
        logging.error("unsupported arguments: %s ", str(unknown))
        sys.exit(8)  # ENOEXEC
    return args


# parse_list - handle the case of list option sanitizing/checking
#
def parse_list(args):
    """Parse arguments in case of list -option/command given."""
    if not os.path.isfile(args.inputfile):
        print("ERROR - input file: '" + args.inputfile + "' does not exist.")
        sys.exit(2)  # ENOENT
    if os.path.isfile(args.listfile + _CSV) or os.path.isfile(
        args.listfile + _XLS
    ):
        if not args.force:
            print("ERROR - output file: '" + args.listfile + _EXISTS)
            sys.exit(2)  # ENOENT
        else:
            print(
                "Warning - output file: '"
                + args.listfile
                + _EXISTS
                + _OVERWRITE
            )
    gen_list(args.inputfile, args.listfile)


# parse_changes - handle the case of changes option sanitizing/checking
#
def parse_changes(args):
    """Parse arguments in case of change -option/command given."""
    if not os.path.isfile(args.previous):
        print("ERROR - previous license file: '" + args.previous + _NOT_EXIST)
        sys.exit(2)  # ENOENT
    if not os.path.isfile(args.current):
        print("ERROR - current license file: '" + args.current + _NOT_EXIST)
        sys.exit(2)  # ENOENT
    if os.path.isfile(args.changefile + _CSV) or os.path.isfile(
        args.changefile + _XLS
    ):
        if not args.force:
            print("ERROR - output file: '" + args.changefile + _EXISTS)
            sys.exit(2)  # ENOENT
        else:
            print(
                "Warning - output file: '"
                + args.changefile
                + _EXISTS
                + _OVERWRITE
            )
    gen_changes(args.previous, args.current, args.changefile)


def main():
    """Script entry point."""
    args = _parse_args()
    if args.command == "list":
        parse_list(args)

    if args.command == "changes":
        parse_changes(args)


if __name__ == "__main__":

    sys.exit(main())
